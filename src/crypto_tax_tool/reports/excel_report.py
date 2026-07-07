from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from crypto_tax_tool.models.fifo_state import AssetLot
from crypto_tax_tool.services.tax_engine import MissingInventoryIssue
from crypto_tax_tool.services.tax_summary import TaxSummary


GERMAN_NUMBER_FORMATS = {"german", "de", "de_DE", "european"}


def _excel_number(value: Decimal) -> float:
    return float(value)


def _number_style(number_format: str, decimals: int = 12) -> str:
    decimal_part = "0" * decimals
    if number_format in GERMAN_NUMBER_FORMATS:
        return f"#.##0,{decimal_part}"
    return f"#,##0.{decimal_part}"


def _eur_style(number_format: str) -> str:
    if number_format in GERMAN_NUMBER_FORMATS:
        return '#.##0,00'
    return '#,##0.00'


class ExcelTaxReportExporter:
    def export(
        self,
        summary: TaxSummary,
        path: Path,
        open_lots: list[AssetLot] | None = None,
        number_format: str = "international",
        missing_inventory_issues: list[MissingInventoryIssue] | None = None,
    ) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tax Summary"

        sheet.append(["Metric", "Value"])
        summary_rows = [
            ["Taxable disposal gain EUR", _excel_number(summary.taxable_gain_eur)],
            ["Long-held disposal gain EUR", _excel_number(summary.long_held_gain_eur)],
            ["Total disposal gain EUR", _excel_number(summary.total_gain_eur)],
            ["Taxable income EUR", _excel_number(summary.taxable_income_eur)],
            ["Total taxable EUR", _excel_number(summary.total_taxable_eur)],
            ["Missing inventory issues", len(missing_inventory_issues or [])],
        ]
        for row in summary_rows:
            sheet.append(row)
        for cell in sheet["B"][1:6]:
            cell.number_format = _eur_style(number_format)

        details = workbook.create_sheet("Disposals")
        details.append(
            [
                "Transaction ID",
                "Asset",
                "Quantity",
                "Proceeds EUR",
                "Cost Basis EUR",
                "Gain EUR",
                "Holding Days Min",
                "Holding Days Max",
                "Classification",
            ]
        )
        for row in summary.rows:
            details.append(
                [
                    row.transaction_id,
                    row.asset,
                    _excel_number(row.quantity),
                    _excel_number(row.proceeds_eur),
                    _excel_number(row.cost_basis_eur),
                    _excel_number(row.gain_eur),
                    row.holding_days_min,
                    row.holding_days_max,
                    row.classification,
                ]
            )
        for row_idx in range(2, details.max_row + 1):
            details.cell(row=row_idx, column=3).number_format = _number_style(number_format)
            for col_idx in (4, 5, 6):
                details.cell(row=row_idx, column=col_idx).number_format = _eur_style(number_format)

        income_sheet = workbook.create_sheet("Taxable Income")
        income_sheet.append(
            [
                "Transaction ID",
                "Received At",
                "Asset",
                "Quantity",
                "Income EUR",
                "EUR Price",
                "Product",
                "Type",
                "Price Provider",
                "Price Pair",
            ]
        )
        for row in summary.income_rows:
            income_sheet.append(
                [
                    row.transaction_id,
                    row.received_at,
                    row.asset,
                    _excel_number(row.quantity),
                    _excel_number(row.income_eur),
                    _excel_number(row.price_eur),
                    row.product,
                    row.raw_type,
                    row.price_provider,
                    row.price_pair,
                ]
            )
        for row_idx in range(2, income_sheet.max_row + 1):
            income_sheet.cell(row=row_idx, column=4).number_format = _number_style(number_format)
            income_sheet.cell(row=row_idx, column=5).number_format = _eur_style(number_format)
            income_sheet.cell(row=row_idx, column=6).number_format = _number_style(number_format)

        lots_sheet = workbook.create_sheet("Open Lots")
        lots_sheet.append(
            [
                "Lot ID",
                "Asset",
                "Acquired At",
                "Original Quantity",
                "Remaining Quantity",
                "Remaining Cost Basis EUR",
                "Source Transaction ID",
            ]
        )
        for lot in open_lots or []:
            lots_sheet.append(
                [
                    lot.id,
                    lot.asset,
                    lot.acquired_at.isoformat(),
                    _excel_number(lot.quantity),
                    _excel_number(lot.remaining_quantity),
                    _excel_number(lot.cost_basis_eur),
                    lot.source_transaction_id,
                ]
            )
        for row_idx in range(2, lots_sheet.max_row + 1):
            lots_sheet.cell(row=row_idx, column=4).number_format = _number_style(number_format)
            lots_sheet.cell(row=row_idx, column=5).number_format = _number_style(number_format)
            lots_sheet.cell(row=row_idx, column=6).number_format = _eur_style(number_format)

        missing_sheet = workbook.create_sheet("Missing Inventory")
        missing_sheet.append(
            [
                "Transaction ID",
                "Disposed At",
                "Asset",
                "Missing Quantity",
                "Disposed Quantity",
                "Proceeds EUR",
                "Product",
                "Type",
                "Message",
                "Suggested Action",
            ]
        )
        for issue in missing_inventory_issues or []:
            missing_sheet.append(
                [
                    issue.transaction_id,
                    issue.disposed_at.isoformat() if issue.disposed_at else "",
                    issue.asset,
                    _excel_number(issue.missing_quantity),
                    _excel_number(issue.disposed_quantity),
                    _excel_number(issue.proceeds_eur),
                    issue.product,
                    issue.raw_type,
                    issue.message,
                    "Add manual FIFO lot or verify missing Binance import; ignore only if immaterial.",
                ]
            )
        for row_idx in range(2, missing_sheet.max_row + 1):
            missing_sheet.cell(row=row_idx, column=4).number_format = _number_style(number_format)
            missing_sheet.cell(row=row_idx, column=5).number_format = _number_style(number_format)
            missing_sheet.cell(row=row_idx, column=6).number_format = _eur_style(number_format)

        workbook.save(path)
        return path
