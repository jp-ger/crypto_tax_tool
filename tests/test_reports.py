from datetime import UTC, datetime
from decimal import Decimal

from openpyxl import load_workbook

from crypto_tax_tool.models.fifo_state import AssetLot, LotUsage
from crypto_tax_tool.reports.csv_report import CsvTaxReportExporter
from crypto_tax_tool.reports.excel_report import ExcelTaxReportExporter
from crypto_tax_tool.services.tax_engine import DisposalResult, TaxCalculationResult
from crypto_tax_tool.services.tax_summary import TaxSummaryService


def _summary():
    disposal = DisposalResult(
        transaction_id="sell1",
        asset="BTC",
        quantity=Decimal("0.1"),
        proceeds_eur=Decimal("5000"),
        matches=[
            LotUsage(
                lot_id="lot1",
                asset="BTC",
                acquired_at=datetime(2025, 1, 1, tzinfo=UTC),
                used_at=datetime(2025, 6, 1, tzinfo=UTC),
                quantity=Decimal("0.1"),
                cost_basis_eur=Decimal("3000"),
                value_eur=Decimal("5000"),
            )
        ],
    )
    return TaxSummaryService().build_summary(
        TaxCalculationResult(disposals=[disposal], open_lots=[])
    )


def test_csv_exporter_creates_file(tmp_path) -> None:
    path = CsvTaxReportExporter().export(_summary(), tmp_path / "report.csv")
    content = path.read_text(encoding="utf-8")
    assert "transaction_id" in content
    assert "sell1" in content


def test_excel_exporter_creates_workbook(tmp_path) -> None:
    open_lot = AssetLot(
        id="lot_open",
        asset="BTC",
        acquired_at=datetime(2025, 1, 1, tzinfo=UTC),
        quantity=Decimal("0.2"),
        remaining_quantity=Decimal("0.1"),
        cost_basis_eur=Decimal("4000"),
        source_transaction_id="buy1",
    )
    path = ExcelTaxReportExporter().export(_summary(), tmp_path / "report.xlsx", [open_lot])
    workbook = load_workbook(path)
    assert "Tax Summary" in workbook.sheetnames
    assert "Disposals" in workbook.sheetnames
    assert "Open Lots" in workbook.sheetnames
    assert workbook["Disposals"][2][0].value == "sell1"
    assert workbook["Open Lots"][2][0].value == "lot_open"
