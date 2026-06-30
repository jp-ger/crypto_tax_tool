from datetime import UTC, datetime
from decimal import Decimal

from openpyxl import load_workbook

from crypto_tax_tool.database.manual_store import save_manual_price
from crypto_tax_tool.database.sqlite_store import initialize_sqlite, save_balance_snapshot, save_transactions
from crypto_tax_tool.models.balances import AssetBalance, BalanceSnapshot
from crypto_tax_tool.models.enums import TaxCategory, TradeSide, TransactionKind, TransactionSource
from crypto_tax_tool.models.manual_entries import ManualPriceEntry
from crypto_tax_tool.models.transactions import NormalizedTransaction
from crypto_tax_tool.services.report_service import ReportGenerationService
from crypto_tax_tool.services.validation_service import ValidationService


def _prepare_db(tmp_path, monkeypatch) -> None:
    monkeypatch.setenv("CRYPTO_TAX_DB_PATH", str(tmp_path / "integration.sqlite3"))

    from crypto_tax_tool.settings import get_settings

    get_settings.cache_clear()
    initialize_sqlite()


def _tx(**kwargs) -> NormalizedTransaction:
    defaults = dict(
        source=TransactionSource.BINANCE,
        source_id="test",
        timestamp=datetime(2025, 1, 1, tzinfo=UTC),
        kind=TransactionKind.TRADE,
        tax_category=TaxCategory.PRIVATE_DISPOSAL,
        asset="BTC",
        quantity=Decimal("0"),
        product="spot",
        raw_type="integration_test",
    )
    defaults.update(kwargs)
    return NormalizedTransaction(**defaults)


def test_end_to_end_report_generation_with_manual_prices(tmp_path, monkeypatch) -> None:
    _prepare_db(tmp_path, monkeypatch)

    buy_time = datetime(2024, 1, 1, tzinfo=UTC)
    sell_time = datetime(2025, 1, 1, tzinfo=UTC)
    save_manual_price(
        ManualPriceEntry(
            asset="USDC",
            quote_asset="EUR",
            timestamp=buy_time,
            price=Decimal("1"),
            reason="Integration test price.",
        )
    )
    save_manual_price(
        ManualPriceEntry(
            asset="USDC",
            quote_asset="EUR",
            timestamp=sell_time,
            price=Decimal("1"),
            reason="Integration test price.",
        )
    )

    buy = _tx(
        source_id="buy_btc_1",
        timestamp=buy_time,
        asset="BTC",
        quantity=Decimal("0.1"),
        quote_asset="USDC",
        quote_quantity=Decimal("3000"),
        side=TradeSide.BUY,
    )
    sell = _tx(
        source_id="sell_btc_1",
        timestamp=sell_time,
        asset="USDC",
        quantity=Decimal("5000"),
        quote_asset="BTC",
        quote_quantity=Decimal("0.1"),
        side=TradeSide.SELL,
    )

    assert save_transactions([buy, sell]) == 2
    save_balance_snapshot(
        BalanceSnapshot(
            source=TransactionSource.BINANCE,
            balances=[AssetBalance(asset="USDC", free=Decimal("2000"), locked=Decimal("0"))],
        )
    )

    validation = ValidationService().validate()
    assert validation.can_create_report

    result = ReportGenerationService().generate_tax_report(output_dir=tmp_path / "report")

    assert result.transactions == 2
    assert result.disposals == 1
    assert result.open_lots == 0
    assert result.summary_csv.exists()
    assert result.summary_xlsx.exists()
    assert result.audit_csv.exists()

    csv_content = result.summary_csv.read_text(encoding="utf-8")
    assert "BTC" in csv_content
    assert "2000" in csv_content

    audit_content = result.audit_csv.read_text(encoding="utf-8")
    assert "BTC" in audit_content

    workbook = load_workbook(result.summary_xlsx)
    assert "Tax Summary" in workbook.sheetnames
    assert "Disposals" in workbook.sheetnames
    assert workbook["Disposals"][2][1].value == "BTC"
